"""Экспорт результатов в Excel."""

from __future__ import annotations

from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from coder.config import ProjectConfig, output_xlsx_path, project_path
from coder.models import Observation, ParsedTranscript, SummaryRow
from coder.prompts import MODALITY_TAGS, PRIMARY_CODES

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _sheet_headers(ws, headers: list[str]) -> None:
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def _autosize(ws) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[letter].width = min(max_len + 2, 60)


def export_workbook(
    project_dir: Path,
    config: ProjectConfig,
    transcripts: list[ParsedTranscript],
    observations: list[Observation],
    summary: list[SummaryRow],
) -> Path:
    wb = Workbook()
    wb.remove(wb.active)

    # Настройки
    ws = wb.create_sheet("Настройки")
    rows = [
        ("Тема", config.research_topic),
        ("Цели", config.research_goals),
        ("Режим", config.run_mode.value),
        ("Респондент (speaker)", config.respondent_speaker),
    ]
    for h in config.hypotheses:
        rows.append(("Гипотеза", h))
    for q in config.research_questions:
        rows.append(("Вопрос", q))
    for t in config.related_themes:
        rows.append(("Связанная тема", t))
    for r in rows:
        ws.append(list(r))

    # Респонденты
    ws = wb.create_sheet("Респонденты")
    field_keys = [f.key for f in config.respondent_fields]
    headers = ["ID", "Файл", "Дата", *field_keys]
    _sheet_headers(ws, headers)
    for r in config.respondents:
        meta = {"segment": r.segment, "interview_date": r.interview_date}
        meta.update(r.extra)
        row = [r.id, r.transcript_file, r.interview_date]
        row.extend(meta.get(k, "") or "" for k in field_keys)
        ws.append(row)
    _autosize(ws)

    # Кодировка
    ws = wb.create_sheet("Кодировка")
    resp_field_labels = [f.label for f in config.respondent_fields]
    resp_field_keys = [f.key for f in config.respondent_fields]
    headers = [
        "ID",
        "Респондент",
        *resp_field_labels,
        "Цитата",
        "Наблюдение",
        "Код",
        "Код 2",
        "Модальность",
        "Severity",
        "Кластер",
        "Категория",
        "JTBD",
        "Workaround",
        "Ключ. задача",
        "Тип",
        "Контекст",
        "Фрагмент",
        "Дата",
    ]
    _sheet_headers(ws, headers)
    for o in sorted(observations, key=lambda x: (x.respondent_id, x.utterance_index or 0)):
        resp_values = [str(o.respondent_meta.get(k, "") or "") for k in resp_field_keys]
        ws.append(
            [
                o.id,
                o.respondent_id,
                *resp_values,
                o.quote,
                o.atom,
                o.primary_code,
                o.secondary_code or "",
                ", ".join(o.modality_tags),
                o.severity_signal,
                o.affinity_cluster,
                o.normalized_category,
                o.jtbd,
                o.workaround,
                "Да" if o.is_key_task else "Нет",
                o.kind,
                o.context,
                o.interview_fragment,
                str(o.respondent_meta.get("interview_date", "") or ""),
            ]
        )
    _autosize(ws)

    # Исключено
    ws = wb.create_sheet("Исключено")
    _sheet_headers(ws, ["Респондент", "Индекс", "Спикер", "Время", "Тип", "Причина", "Текст"])
    for tr in transcripts:
        for u in tr.utterances:
            if u.include_in_coding:
                continue
            ws.append(
                [
                    tr.respondent_id,
                    u.index,
                    u.speaker,
                    u.timestamp_raw,
                    u.segment_type.value if u.segment_type else "",
                    u.exclude_reason,
                    u.text[:500],
                ]
            )
    _autosize(ws)

    # Транскрипты
    ws = wb.create_sheet("Транскрипты")
    _sheet_headers(ws, ["Респондент", "Время", "Спикер", "В кодировке", "Текст", "ID наблюдений"])
    obs_by_utt: dict[tuple[str, int], list[str]] = {}
    for o in observations:
        if o.utterance_index is not None:
            obs_by_utt.setdefault((o.respondent_id, o.utterance_index), []).append(o.id)
    for tr in transcripts:
        for u in tr.utterances:
            ids = ", ".join(obs_by_utt.get((tr.respondent_id, u.index), []))
            ws.append(
                [
                    tr.respondent_id,
                    u.timestamp_raw,
                    u.speaker,
                    "Да" if u.include_in_coding else "Нет",
                    u.text[:800],
                    ids,
                ]
            )
    _autosize(ws)

    # Сводная
    ws = wb.create_sheet("Сводная")

    # Собираем наблюдения по кластеру для агрегации мета-полей
    obs_by_cluster: dict[str, list[Observation]] = defaultdict(list)
    for o in observations:
        obs_by_cluster[o.affinity_cluster or "Без кластера"].append(o)

    # Динамические столбцы: все кастомные поля респондентов
    meta_fields = config.respondent_fields  # RespondentField list
    meta_headers = [f.label for f in meta_fields]
    meta_keys = [f.key for f in meta_fields]

    _sheet_headers(
        ws,
        [
            "Кластер", "Категория", "Описание",
            "Кол-во респондентов", "Респонденты", "Частота",
            "Коды", "Модальности",
            *meta_headers,
            "Критичность", "Цитаты", "ID наблюдений",
        ],
    )
    for s in summary:
        cluster_obs = obs_by_cluster.get(s.affinity_cluster, [])

        # Уникальные значения каждого мета-поля для кластера
        meta_values = []
        for key in meta_keys:
            vals = ", ".join(sorted({
                str(o.respondent_meta.get(key, "") or "")
                for o in cluster_obs
                if o.respondent_meta.get(key)
            }))
            meta_values.append(vals)

        ws.append(
            [
                s.affinity_cluster,
                s.normalized_category,
                s.description,
                s.respondent_count,
                s.respondent_ids_str,
                s.cluster_frequency,
                s.all_codes,
                s.dominant_modalities,
                *meta_values,
                s.criticality,
                s.representative_quotes,
                ", ".join(s.observation_ids),
            ]
        )
    _autosize(ws)

    if config.run_mode.value == "test":
        ws = wb.create_sheet("Тест")
        ws.append(["Тестовый прогон", f"scope: {config.test_scope.type.value} = {config.test_scope.value}"])
        ws.append(["Наблюдений", len(observations)])

    # Справочник / Codebook
    ws = wb.create_sheet("Справочник")
    _sheet_headers(ws, ["Тип", "Код / Тег", "Определение", "Когда использовать", "Когда НЕ использовать", "Речевые маркеры"])
    codebook_rows = [
        ("Behavioral code", "Задача", "Цель или намерение пользователя", "Что пользователь хочет сделать/достичь", "Если это устойчивый паттерн, а не разовое намерение → Паттерн", ""),
        ("Behavioral code", "Триггер", "Событие, запускающее немедленное действие", "Внешний стимул → немедленная реакция", "Если действие регулярное без конкретного события → Паттерн", ""),
        ("Behavioral code", "Паттерн поведения", "Устойчивый, повторяющийся способ действия", "Обычный, привычный способ работы", "Если это разовое намерение → Задача", "«всегда», «обычно», «как правило»"),
        ("Behavioral code", "Драйвер", "Фактор, ускоряющий или упрощающий задачу", "Что помогает пользователю выполнить задачу", "Если пользователь осознаёт ценность → можно добавить Ценность вторым кодом", ""),
        ("Behavioral code", "Проблема", "Трудность, замедляющая задачу; пользователь всё равно её решает", "Задача выполнена, но с трудом или обходным путём", "Если пользователь ОТКАЗАЛСЯ или ИЗБЕЖАЛ задачи → Барьер", "«раздражает», «неудобно», «приходится», «каждый раз заново»"),
        ("Behavioral code", "Барьер", "Фактор, ведущий к отказу, избеганию или откладыванию задачи", "Задача НЕ выполнена или избегается", "Если пользователь всё-таки выполнил задачу → Проблема", "«ушёл», «не стал», «бросил», «невозможно»"),
        ("Behavioral code", "Ценность", "Осознаваемая польза или выгода от продукта/действия", "Пользователь явно говорит о пользе", "Если это просто факт без оценки полезности", "«помогает», «удобно», «нравится», «ценно»"),
        ("Behavioral code", "Незнание возможности", "Пользователь не знает о существующей функции", "Пользователь обходит проблему, не зная о решении", "Если пользователь знает, но не использует → Паттерн или Барьер", "«не знал», «оказывается можно», «а разве есть...»"),
        ("Behavioral code", "Недоверие / сомнение", "Скептическое отношение к продукту, данным, обещаниям", "Явное недоверие или сомнение", "Если это просто вопрос без негативного отношения", "«не уверен», «вдруг», «боюсь», «не верю»"),
        ("Modality", "Необходимость", "Пользователь вынужден делать что-то", "Явный маркер вынужденности", "", "«приходится», «вынужден», «обязан», «надо»"),
        ("Modality", "Невозможность", "Пользователь не может выполнить задачу", "Явный маркер невозможности", "", "«не могу», «невозможно», «не получается», «не даёт»"),
        ("Modality", "Возможность", "Может, но не всегда", "Условная или частичная возможность", "", "«можно было бы», «иногда получается»"),
        ("Modality", "Уверенность", "Твёрдое убеждение", "Явный маркер уверенности", "", "«точно», «уверен», «однозначно»"),
        ("Modality", "Сомнение", "Колебание, неуверенность", "Явный маркер неуверенности", "", "«вроде», «может быть», «не уверен»"),
        ("Modality", "Оценка+", "Положительное отношение", "Явный позитивный маркер", "", "«круто», «удобно», «нравится», «здорово»"),
        ("Modality", "Оценка−", "Отрицательное отношение", "Явный негативный маркер", "", "«раздражает», «ужасно», «бесит», «неудобно»"),
        ("Modality", "Норма", "Ожидаемое, обычное", "Явный маркер нормы", "", "«как обычно», «так и должно быть», «всегда так»"),
        ("Modality", "Желательность", "Хотение, предпочтение", "Явный маркер желания", "", "«хочу», «хотелось бы», «было бы здорово»"),
        ("Modality", "Гипотетичность", "Предположение, условие", "Явный маркер гипотезы", "", "«если бы», «в случае если», «представьте»"),
    ]
    for row in codebook_rows:
        ws.append(list(row))
    _autosize(ws)

    # Цитаты / Quote Library
    ws = wb.create_sheet("Цитаты")
    _sheet_headers(ws, ["Цитата", "Наблюдение", "Кластер", "Категория", "Код", "Модальность", "Severity", "Респондент"])
    for o in sorted(observations, key=lambda x: x.affinity_cluster):
        if not o.quote or len(o.quote.strip()) < 20:
            continue
        ws.append([
            o.quote,
            o.atom,
            o.affinity_cluster,
            o.normalized_category,
            o.primary_code,
            ", ".join(o.modality_tags),
            o.severity_signal,
            o.respondent_id,
        ])
    _autosize(ws)

    path = output_xlsx_path(project_path(project_dir), config)
    wb.save(path)
    return path
